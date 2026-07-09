from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import re
import uuid
import jwt
from io import BytesIO
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
from datetime import datetime, timezone, timedelta, date
from pwdlib import PasswordHash
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
TOKEN_EXPIRE_DAYS = 7

pwd_hash = PasswordHash.recommended()
security = HTTPBearer()

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ------------------- Constants -------------------
ROLES = [
    "marketing_head", "marketing_deputy_head", "product_executive",
    "general_manager", "ceo", "chairman", "agm", "business_manager",
]
ADMIN_ROLES = {"marketing_head", "marketing_deputy_head", "chairman"}
USER_MANAGER_ROLES = {"chairman"}
CATEGORIES = ["task", "sales_collection", "target"]
STATUSES = ["pending", "in_progress", "completed"]
HEADS = ["company", "scientific_inputs", "engagement"]


# ------------------- Models -------------------
class RegisterRequest(BaseModel):
    name: str
    email: EmailStr
    password: str = Field(min_length=6)
    role: str
    hq: Optional[str] = None


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class UserPublic(BaseModel):
    id: str
    name: str
    email: str
    role: str
    hq: Optional[str] = None


class TokenResponse(BaseModel):
    access_token: str
    user: UserPublic


class TaskCreate(BaseModel):
    title: str
    description: Optional[str] = None
    assignee: Optional[str] = None
    role: Optional[str] = None
    hq: Optional[str] = None
    frequency: str = "weekly"  # daily | weekly
    category: str = "task"  # task | sales_collection | target
    head: Optional[str] = None  # company | scientific_inputs | engagement
    start_date: Optional[str] = None  # YYYY-MM-DD
    due_date: Optional[str] = None
    reporting_due_date: Optional[str] = None
    target_amount: Optional[float] = None


class TaskUpdate(BaseModel):
    status: Optional[str] = None
    collected_amount: Optional[float] = None
    title: Optional[str] = None
    description: Optional[str] = None
    due_date: Optional[str] = None


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str = Field(min_length=6)


class AdminUserUpdate(BaseModel):
    role: Optional[str] = None
    hq: Optional[str] = None
    name: Optional[str] = None


class AdminResetPasswordRequest(BaseModel):
    new_password: str = Field(min_length=6)


class PhotoUpload(BaseModel):
    photo_base64: str
    caption: Optional[str] = None


# ------------------- Auth helpers -------------------
def create_token(user: dict) -> str:
    payload = {
        "sub": user["id"],
        "email": user["email"],
        "role": user["role"],
        "exp": datetime.now(timezone.utc) + timedelta(days=TOKEN_EXPIRE_DAYS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    user = await db.users.find_one({"id": payload.get("sub")}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] not in ADMIN_ROLES:
        raise HTTPException(status_code=403, detail="Only Marketing Head / Deputy Head / Chairman can perform this action")
    return user


async def require_user_manager(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] not in USER_MANAGER_ROLES:
        raise HTTPException(status_code=403, detail="Only Chairman can manage users")
    return user


# ------------------- Date helpers -------------------
def utc_today() -> date:
    return datetime.now(timezone.utc).date()


def period_range(period: str):
    t = utc_today()
    if period == "week":
        start = t - timedelta(days=t.weekday())
        end = start + timedelta(days=6)
    elif period == "quarter":
        q = (t.month - 1) // 3
        start = date(t.year, q * 3 + 1, 1)
        end_month = q * 3 + 3
        if end_month == 12:
            end = date(t.year, 12, 31)
        else:
            end = date(t.year, end_month + 1, 1) - timedelta(days=1)
    else:  # month
        start = date(t.year, t.month, 1)
        if t.month == 12:
            end = date(t.year, 12, 31)
        else:
            end = date(t.year, t.month + 1, 1) - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def parse_excel_date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%m/%d/%Y", "%d %b %Y", "%d-%b-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ------------------- Excel header mapping -------------------
HEADER_MAP = {
    "task name": "title", "task": "title", "title": "title", "task title": "title", "activity": "title",
    "description": "description", "details": "description", "task description": "description",
    "assignee": "assignee", "assigned to": "assignee", "assignee name": "assignee", "responsible person": "assignee",
    "role": "role", "roles and responsibilities": "role", "responsibility": "role", "designation": "role",
    "hq": "hq", "headquarter": "hq", "headquarters": "hq", "hq name": "hq",
    "frequency": "frequency", "freq": "frequency", "task frequency": "frequency",
    "start date": "start_date", "from date": "start_date",
    "due date": "due_date", "end date": "due_date", "deadline": "due_date", "timeline": "due_date", "to date": "due_date",
    "category": "category", "type": "category", "task type": "category",
    "head": "head", "activity head": "head", "activity type": "head", "segment": "head", "activity segment": "head",
    "target amount": "target_amount", "target": "target_amount", "sales target": "target_amount",
    "collected amount": "collected_amount", "collection": "collected_amount", "sales collection": "collected_amount", "collection amount": "collected_amount",
    "reporting due date": "reporting_due_date", "reporting date": "reporting_due_date",
    "report due date": "reporting_due_date", "due date of reporting": "reporting_due_date",
}


def norm_header(h) -> str:
    return str(h).strip().lower().replace("_", " ").replace("*", "").strip()


def norm_category(v) -> str:
    if not v:
        return "task"
    s = str(v).strip().lower()
    if "collection" in s:
        return "sales_collection"
    if "target" in s:
        return "target"
    return "task"


def norm_frequency(v) -> str:
    if not v:
        return "weekly"
    s = str(v).strip().lower()
    return "daily" if s.startswith("d") else "weekly"


def norm_head(v) -> Optional[str]:
    if not v:
        return None
    s = str(v).strip().lower()
    if "scien" in s or "input" in s:
        return "scientific_inputs"
    if "engag" in s:
        return "engagement"
    if "comp" in s:
        return "company"
    return None


def to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("₹", "").strip())
    except ValueError:
        return None


def task_doc(data: dict, created_by: str) -> dict:
    now = datetime.now(timezone.utc).isoformat()
    due = data.get("due_date")
    return {
        "id": str(uuid.uuid4()),
        "title": data.get("title") or "",
        "description": data.get("description"),
        "assignee": data.get("assignee"),
        "role": data.get("role"),
        "hq": data.get("hq"),
        "frequency": data.get("frequency") or "weekly",
        "category": data.get("category") or "task",
        "head": data.get("head"),
        "start_date": data.get("start_date"),
        "due_date": due,
        "reporting_due_date": data.get("reporting_due_date"),
        "target_amount": data.get("target_amount"),
        "collected_amount": data.get("collected_amount"),
        "status": "pending",
        "month": due[:7] if due else None,
        "created_by": created_by,
        "created_at": now,
        "updated_at": now,
    }


# ------------------- Auth routes -------------------
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(req: RegisterRequest):
    if req.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    existing = await db.users.find_one({"email": req.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="An account with this email already exists")
    user = {
        "id": str(uuid.uuid4()),
        "name": req.name.strip(),
        "email": req.email.lower(),
        "role": req.role,
        "hq": req.hq,
        "password_hash": pwd_hash.hash(req.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(dict(user))
    public = UserPublic(**{k: user[k] for k in ("id", "name", "email", "role", "hq")})
    return TokenResponse(access_token=create_token(user), user=public)


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(req: LoginRequest):
    user = await db.users.find_one({"email": req.email.lower()}, {"_id": 0})
    if not user or not pwd_hash.verify(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    public = UserPublic(**{k: user.get(k) for k in ("id", "name", "email", "role", "hq")})
    return TokenResponse(access_token=create_token(user), user=public)


@api_router.get("/auth/me", response_model=UserPublic)
async def me(user: dict = Depends(get_current_user)):
    return UserPublic(**{k: user.get(k) for k in ("id", "name", "email", "role", "hq")})


@api_router.post("/auth/change-password")
async def change_password(req: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    doc = await db.users.find_one({"id": user["id"]})
    if not doc or not pwd_hash.verify(req.current_password, doc["password_hash"]):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"password_hash": pwd_hash.hash(req.new_password)}},
    )
    return {"detail": "Password updated successfully"}


@api_router.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(500)
    return users


# ------------------- User management (Chairman only) -------------------
@api_router.patch("/admin/users/{user_id}")
async def admin_update_user(user_id: str, req: AdminUserUpdate, admin: dict = Depends(require_user_manager)):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    if "role" in updates and updates["role"] not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    if updates:
        await db.users.update_one({"id": user_id}, {"$set": updates})
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated


@api_router.post("/admin/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, req: AdminResetPasswordRequest, admin: dict = Depends(require_user_manager)):
    target = await db.users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": pwd_hash.hash(req.new_password)}},
    )
    return {"detail": "Password reset successfully"}


@api_router.delete("/admin/users/{user_id}")
async def admin_delete_user(user_id: str, admin: dict = Depends(require_user_manager)):
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"deleted": True}


# ------------------- Task routes -------------------
@api_router.post("/tasks")
async def create_task(req: TaskCreate, user: dict = Depends(get_current_user)):
    data = req.model_dump()
    data["category"] = norm_category(data.get("category"))
    data["frequency"] = norm_frequency(data.get("frequency"))
    data["head"] = norm_head(data.get("head"))
    doc = task_doc(data, user["id"])
    if not doc["title"].strip():
        raise HTTPException(status_code=400, detail="Title is required")
    await db.tasks.insert_one(dict(doc))
    return doc


@api_router.get("/tasks")
async def list_tasks(
    frequency: Optional[str] = None,
    hq: Optional[str] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    head: Optional[str] = None,
    assignee: Optional[str] = None,
    role: Optional[str] = None,
    period: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    query = {}
    if frequency:
        query["frequency"] = frequency
    if hq:
        query["hq"] = hq
    if status:
        query["status"] = status
    if category:
        query["category"] = category
    if head:
        query["head"] = head
    if assignee:
        query["assignee"] = assignee
    if role:
        query["role"] = role
    if period in ("week", "month", "quarter"):
        start, end = period_range(period)
        query["due_date"] = {"$gte": start, "$lte": end}
    if search:
        query["title"] = {"$regex": search, "$options": "i"}
    pipeline = [
        {"$match": query},
        {"$sort": {"due_date": 1}},
        {"$limit": 1000},
        {"$addFields": {"photo_count": {"$size": {"$ifNull": ["$photos", []]}}}},
        {"$project": {"_id": 0, "photos": 0}},
    ]
    tasks = await db.tasks.aggregate(pipeline).to_list(1000)
    return tasks


@api_router.get("/tasks/{task_id}")
async def get_task(task_id: str, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@api_router.patch("/tasks/{task_id}")
async def update_task(task_id: str, req: TaskUpdate, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    if "status" in updates and updates["status"] not in STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if "status" in updates and updates["status"] == "completed":
        updates["completed_at"] = datetime.now(timezone.utc).isoformat()
    if "due_date" in updates:
        updates["month"] = updates["due_date"][:7]
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.tasks.update_one({"id": task_id}, {"$set": updates})
    updated = await db.tasks.find_one({"id": task_id}, {"_id": 0})
    return updated


@api_router.delete("/tasks/{task_id}")
async def delete_task(task_id: str, user: dict = Depends(require_admin)):
    result = await db.tasks.delete_one({"id": task_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"deleted": True}


# ------------------- Task photos (marketing activity proof) -------------------
@api_router.post("/tasks/{task_id}/photos")
async def add_task_photo(task_id: str, req: PhotoUpload, user: dict = Depends(get_current_user)):
    task = await db.tasks.find_one({"id": task_id}, {"_id": 0, "id": 1, "photos": 1})
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if len(req.photo_base64) > 4_000_000:
        raise HTTPException(status_code=400, detail="Photo is too large. Please retake with lower quality.")
    if len(task.get("photos") or []) >= 10:
        raise HTTPException(status_code=400, detail="Maximum 10 photos per task")
    photo = {
        "id": str(uuid.uuid4()),
        "data": req.photo_base64,
        "caption": req.caption,
        "uploaded_by": user["name"],
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.tasks.update_one({"id": task_id}, {"$push": {"photos": photo}})
    return photo


@api_router.delete("/tasks/{task_id}/photos/{photo_id}")
async def delete_task_photo(task_id: str, photo_id: str, user: dict = Depends(get_current_user)):
    result = await db.tasks.update_one({"id": task_id}, {"$pull": {"photos": {"id": photo_id}}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Task not found")
    return {"deleted": True}


# ------------------- Excel upload -------------------
@api_router.post("/tasks/upload")
async def upload_excel(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx Excel files are supported")
    contents = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Please check the format.")
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    headers = rows[0]
    col_map = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        key = HEADER_MAP.get(norm_header(h))
        if key and key not in col_map:
            col_map[key] = idx
    if "title" not in col_map:
        raise HTTPException(status_code=400, detail="Could not find a 'Task Name' column in the Excel file")

    inserted = []
    skipped = []
    for row_num, row in enumerate(rows[1:], start=2):
        def cell(key):
            i = col_map.get(key)
            return row[i] if i is not None and i < len(row) else None

        title = cell("title")
        if title is None or str(title).strip() == "":
            if any(v is not None and str(v).strip() != "" for v in row):
                skipped.append({"row": row_num, "reason": "Missing task name"})
            continue
        data = {
            "title": str(title).strip(),
            "description": str(cell("description")).strip() if cell("description") is not None else None,
            "assignee": str(cell("assignee")).strip() if cell("assignee") is not None else None,
            "role": str(cell("role")).strip() if cell("role") is not None else None,
            "hq": str(cell("hq")).strip() if cell("hq") is not None else None,
            "frequency": norm_frequency(cell("frequency")),
            "category": norm_category(cell("category")),
            "head": norm_head(cell("head")),
            "start_date": parse_excel_date(cell("start_date")),
            "due_date": parse_excel_date(cell("due_date")),
            "reporting_due_date": parse_excel_date(cell("reporting_due_date")),
            "target_amount": to_float(cell("target_amount")),
            "collected_amount": to_float(cell("collected_amount")),
        }
        inserted.append(task_doc(data, user["id"]))

    if inserted:
        await db.tasks.insert_many([dict(d) for d in inserted])
    return {"inserted_count": len(inserted), "skipped": skipped, "filename": file.filename}


# ------------------- Performance sheets (Brand / Territory / Management) -------------------
MONTH_MAP = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
             "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def parse_sheet_month(sheet_name: str) -> Optional[str]:
    n = sheet_name.lower()
    month = None
    for abbr, num in MONTH_MAP.items():
        if abbr in n:
            month = num
            break
    if month is None:
        return None
    year = utc_today().year
    m4 = re.search(r"20\d{2}", n)
    if m4:
        year = int(m4.group(0))
    else:
        m2 = re.search(r"\b(\d{2})\b", n)
        if m2:
            year = 2000 + int(m2.group(1))
    return f"{year}-{month:02d}"


def sheet_text(ws, max_row=8) -> str:
    parts = []
    for row in ws.iter_rows(max_row=max_row, values_only=True):
        for c in row:
            if c is not None:
                parts.append(str(c))
    return " ".join(parts).lower()


def detect_perf_type(wb) -> Optional[str]:
    for sn in wb.sheetnames:
        text = sheet_text(wb[sn])
        if "total primary sales" in text or "active doctors" in text:
            return "management"
        if "h.q" in text or "be/kam" in text or "territory management" in text or "sec status" in text:
            return "territory"
        if "brand" in text and "target" in text:
            return "brand"
    return None


def parse_brand_sheet(ws, month: str) -> List[dict]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] is not None and str(row[0]).strip().lower() == "brand":
            header_idx = i
            break
    if header_idx is None:
        return []
    docs = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        brand = str(row[0]).strip()
        if not brand or brand.lower() == "total":
            continue

        def g(i):
            return row[i] if i < len(row) else None

        target = to_float(g(1))
        weeks = [to_float(g(i)) for i in (2, 3, 4, 5)]
        final = next((w for w in reversed(weeks) if w is not None), None)
        ach = round(final / target * 100, 1) if (target and final is not None) else None
        docs.append({
            "id": str(uuid.uuid4()),
            "month": month,
            "brand": brand,
            "target": target,
            "w1": weeks[0], "w2": weeks[1], "w3": weeks[2], "w4": weeks[3],
            "sales_total": final,
            "achievement_pct": ach,
            "growth": str(g(6)).strip() if g(6) is not None else None,
            "top_territory": str(g(7)).strip() if g(7) is not None else None,
            "low_territory": str(g(8)).strip() if g(8) is not None else None,
        })
    return docs


MGMT_METRICS = [
    ("total primary sales", "primary_sales", "numeric"),
    ("total secondary sales", "secondary_sales", "numeric"),
    ("monthly run rate", "run_rate", "numeric"),
    ("active doctors", "active_doctors", "numeric"),
    ("new prescribers", "new_prescribers", "numeric"),
    ("top brand", "top_brand", "text"),
    ("lowest brand", "lowest_brand", "text"),
    ("strong territory", "strong_territory", "text"),
    ("weak territory", "weak_territory", "text"),
]


def parse_mgmt_sheet(ws, month: str) -> Optional[dict]:
    metrics = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().lower()
        for match, key, kind in MGMT_METRICS:
            if match in label:
                def g(i):
                    return row[i] if i < len(row) else None
                if kind == "numeric":
                    metrics[key] = {
                        "weeks": [to_float(g(i)) for i in (1, 2, 3, 4)],
                        "total": to_float(g(5)),
                    }
                else:
                    metrics[key] = {
                        "weeks": [str(g(i)).strip() if g(i) is not None else None for i in (1, 2, 3, 4)],
                    }
                break
    if not metrics:
        return None
    return {"id": str(uuid.uuid4()), "month": month, "metrics": metrics}


TERR_SKIP_PREFIXES = ("h.q", "region / hq", "territory management", "total")


def parse_territory_sheet(ws, month: str) -> List[dict]:
    docs = []
    current_region = None
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        s0 = str(row[0]).strip()
        low = s0.lower()
        if not s0:
            continue

        def g(i):
            return row[i] if i < len(row) else None

        if "region" in low and g(1) is None and g(3) is None:
            current_region = re.sub(r"\s*\(.*\)\s*$", "", s0).strip()
            continue
        if any(low.startswith(p) for p in TERR_SKIP_PREFIXES) or "sec status" in low:
            continue
        # data row must have a person name or a target/sales value
        if g(1) is None and g(3) is None and g(7) is None:
            continue
        target = to_float(g(3))
        weeks = [to_float(g(i)) for i in (4, 5, 6, 7)]
        final = next((w for w in reversed(weeks) if w is not None), None)
        ach = to_float(g(8))
        if ach is None and target and final is not None:
            ach = final / target * 100
        docs.append({
            "id": str(uuid.uuid4()),
            "month": month,
            "region": current_region or "Other",
            "hq": s0,
            "be_name": str(g(1)).strip() if g(1) is not None else None,
            "doj": parse_excel_date(g(2)),
            "target": target,
            "w1": weeks[0], "w2": weeks[1], "w3": weeks[2], "w4": weeks[3],
            "sales_total": final,
            "achievement_pct": round(ach, 1) if ach is not None else None,
        })
    return docs


PERF_COLLECTIONS = {"brand": "brand_performance", "territory": "territory_performance",
                    "management": "management_dashboard"}


@api_router.post("/performance/upload")
async def upload_performance(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx Excel files are supported")
    contents = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Please check the format.")

    ptype = detect_perf_type(wb)
    if ptype is None:
        raise HTTPException(
            status_code=400,
            detail="Could not recognize this sheet. Supported: Brand Performance, Territory Performance, Management Dashboard, or the monthly Task sheet (use Task upload for that).",
        )

    collection = db[PERF_COLLECTIONS[ptype]]
    now = datetime.now(timezone.utc).isoformat()
    months_parsed = []
    total_inserted = 0
    for sn in wb.sheetnames:
        month = parse_sheet_month(sn)
        if month is None:
            continue
        ws = wb[sn]
        if ptype == "brand":
            docs = parse_brand_sheet(ws, month)
        elif ptype == "territory":
            docs = parse_territory_sheet(ws, month)
        else:
            doc = parse_mgmt_sheet(ws, month)
            docs = [doc] if doc else []
        if not docs:
            continue
        for d in docs:
            d["created_at"] = now
        await collection.delete_many({"month": month})
        await collection.insert_many([dict(d) for d in docs])
        months_parsed.append(month)
        total_inserted += len(docs)

    if total_inserted == 0:
        raise HTTPException(status_code=400, detail="No data rows could be parsed from this file")
    return {"type": ptype, "months": sorted(set(months_parsed)), "inserted_count": total_inserted,
            "filename": file.filename}


@api_router.get("/performance/months")
async def performance_months(user: dict = Depends(get_current_user)):
    out = {}
    all_months = set()
    for key, coll in PERF_COLLECTIONS.items():
        months = await db[coll].distinct("month")
        out[key] = sorted(months, reverse=True)
        all_months.update(months)
    out["all"] = sorted(all_months, reverse=True)
    return out


@api_router.get("/performance/brands")
async def performance_brands(month: Optional[str] = None, user: dict = Depends(get_current_user)):
    if not month:
        months = await db.brand_performance.distinct("month")
        if not months:
            return {"month": None, "items": []}
        month = sorted(months, reverse=True)[0]
    items = await db.brand_performance.find({"month": month}, {"_id": 0}).to_list(500)
    items.sort(key=lambda x: -(x.get("sales_total") or 0))
    return {"month": month, "items": items}


@api_router.get("/performance/territories")
async def performance_territories(month: Optional[str] = None, user: dict = Depends(get_current_user)):
    if not month:
        months = await db.territory_performance.distinct("month")
        if not months:
            return {"month": None, "regions": []}
        month = sorted(months, reverse=True)[0]
    items = await db.territory_performance.find({"month": month}, {"_id": 0}).to_list(1000)
    regions = {}
    for it in items:
        regions.setdefault(it.get("region") or "Other", []).append(it)
    out = []
    for region, its in regions.items():
        target = sum(i.get("target") or 0 for i in its)
        sales = sum(i.get("sales_total") or 0 for i in its)
        out.append({
            "region": region,
            "items": its,
            "target_total": round(target, 2),
            "sales_total": round(sales, 2),
            "achievement_pct": round(sales / target * 100, 1) if target else None,
        })
    out.sort(key=lambda r: -r["sales_total"])
    return {"month": month, "regions": out}


@api_router.get("/performance/management")
async def performance_management(month: Optional[str] = None, user: dict = Depends(get_current_user)):
    if not month:
        months = await db.management_dashboard.distinct("month")
        if not months:
            return {"month": None, "metrics": None}
        month = sorted(months, reverse=True)[0]
    doc = await db.management_dashboard.find_one({"month": month}, {"_id": 0})
    return {"month": month, "metrics": (doc or {}).get("metrics")}


async def compute_brand_growth() -> dict:
    items = await db.brand_performance.find({}, {"_id": 0}).to_list(2000)
    months = sorted({i["month"] for i in items})
    brands = {}
    for i in items:
        key = i["brand"].strip().upper()
        brands.setdefault(key, {"brand": i["brand"].strip(), "by_month": {}})["by_month"][i["month"]] = i
    out = []
    for b in brands.values():
        series = []
        prev = None
        for m in months:
            it = b["by_month"].get(m)
            sales = it.get("sales_total") if it else None
            growth = None
            if sales is not None and prev not in (None, 0):
                growth = round((sales - prev) / prev * 100, 1)
            series.append({
                "month": m,
                "sales": sales,
                "target": it.get("target") if it else None,
                "growth_pct": growth,
            })
            if sales is not None:
                prev = sales
        vals = [s["sales"] for s in series if s["sales"] is not None]
        overall = round((vals[-1] - vals[0]) / vals[0] * 100, 1) if len(vals) >= 2 and vals[0] else None
        out.append({
            "brand": b["brand"],
            "series": series,
            "overall_growth_pct": overall,
            "latest_sales": vals[-1] if vals else None,
        })
    out.sort(key=lambda x: -(x["latest_sales"] or 0))
    return {"months": months, "brands": out}


@api_router.get("/performance/growth")
async def performance_growth(user: dict = Depends(get_current_user)):
    return await compute_brand_growth()


# ------------------- Meta / filters -------------------
@api_router.get("/meta/filters")
async def meta_filters(user: dict = Depends(get_current_user)):
    hqs = [h for h in await db.tasks.distinct("hq") if h]
    assignees = [a for a in await db.tasks.distinct("assignee") if a]
    roles = [r for r in await db.tasks.distinct("role") if r]
    return {"hqs": sorted(hqs), "assignees": sorted(assignees), "roles": sorted(roles),
            "categories": CATEGORIES, "statuses": STATUSES, "heads": HEADS}


# ------------------- Dashboard & Reports -------------------
def summarize(tasks: List[dict]):
    total = len(tasks)
    completed = sum(1 for t in tasks if t["status"] == "completed")
    in_progress = sum(1 for t in tasks if t["status"] == "in_progress")
    pending = sum(1 for t in tasks if t["status"] == "pending")
    today = utc_today().isoformat()
    overdue = sum(1 for t in tasks if t.get("due_date") and t["due_date"] < today and t["status"] != "completed")
    rate = round(completed / total * 100, 1) if total else 0.0
    return {"total": total, "completed": completed, "in_progress": in_progress,
            "pending": pending, "overdue": overdue, "completion_rate": rate}


def group_summary(tasks: List[dict], key: str):
    groups = {}
    for t in tasks:
        k = t.get(key) or "Unassigned"
        groups.setdefault(k, []).append(t)
    out = []
    for name, items in groups.items():
        s = summarize(items)
        out.append({"name": name, **s})
    out.sort(key=lambda x: (-x["total"], x["name"]))
    return out


def sales_summary(tasks: List[dict]):
    sales_tasks = [t for t in tasks if t["category"] in ("sales_collection", "target")]
    target_total = sum(t.get("target_amount") or 0 for t in sales_tasks)
    collected_total = sum(t.get("collected_amount") or 0 for t in sales_tasks)
    pct = round(collected_total / target_total * 100, 1) if target_total else 0.0
    return {"target_total": target_total, "collected_total": collected_total,
            "achievement_pct": pct, "count": len(sales_tasks)}


@api_router.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    start, end = period_range("month")
    month_tasks = await db.tasks.find(
        {"due_date": {"$gte": start, "$lte": end}}, {"_id": 0, "photos": 0}).to_list(2000)
    today = utc_today().isoformat()
    todays = [t for t in month_tasks if t.get("due_date") == today or
              (t["frequency"] == "daily" and t["status"] != "completed")][:10]
    recent = await db.tasks.find({}, {"_id": 0, "photos": 0}).sort("created_at", -1).to_list(5)
    return {
        "kpis": summarize(month_tasks),
        "sales": sales_summary(month_tasks),
        "todays_tasks": todays,
        "recent_tasks": recent,
        "period": {"start": start, "end": end},
    }


HEAD_LABELS = {"company": "Company", "scientific_inputs": "Scientific Inputs",
               "engagement": "Engagement", "Unassigned": "Unassigned"}


async def build_report(period: str) -> dict:
    if period not in ("week", "month", "quarter"):
        period = "month"
    start, end = period_range(period)
    tasks = await db.tasks.find(
        {"due_date": {"$gte": start, "$lte": end}}, {"_id": 0, "photos": 0}).to_list(5000)
    return {
        "period": period,
        "range": {"start": start, "end": end},
        "kpis": summarize(tasks),
        "by_hq": group_summary(tasks, "hq"),
        "by_assignee": group_summary(tasks, "assignee"),
        "by_role": group_summary(tasks, "role"),
        "by_frequency": group_summary(tasks, "frequency"),
        "by_head": group_summary(tasks, "head"),
        "sales": sales_summary(tasks),
    }


@api_router.get("/reports")
async def reports(period: str = "month", user: dict = Depends(get_current_user)):
    return await build_report(period)


@api_router.get("/reports/pdf")
async def reports_pdf(period: str = "month", user: dict = Depends(get_current_user)):
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    data = await build_report(period)
    kpis = data["kpis"]
    sales = data["sales"]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm, topMargin=16 * mm, bottomMargin=16 * mm,
        title=f"Surishi Marketing Report - {period.title()}",
    )
    styles = getSampleStyleSheet()
    brand = rl.HexColor("#0055A4")
    gold = rl.HexColor("#C98F12")
    title_style = ParagraphStyle("T", parent=styles["Title"], textColor=brand, fontSize=18, spaceAfter=2)
    sub_style = ParagraphStyle("S", parent=styles["Normal"], textColor=rl.HexColor("#64748B"), fontSize=10)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=brand, fontSize=13, spaceBefore=14, spaceAfter=6)

    period_label = {"week": "Weekly", "month": "Monthly", "quarter": "Quarterly"}[data["period"]]
    story = [
        Paragraph("Surishi Pharmaceuticals", title_style),
        Paragraph(f"Marketing Execution — {period_label} Report", styles["Heading3"]),
        Paragraph(
            f"Period: {data['range']['start']} to {data['range']['end']} &nbsp;·&nbsp; "
            f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y, %H:%M UTC')} &nbsp;·&nbsp; surishi.in",
            sub_style,
        ),
        Spacer(1, 10),
    ]

    def styled_table(rows, col_widths=None, header_bg=brand):
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, rl.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl.white, rl.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        return t

    story.append(Paragraph("Key Metrics", h2))
    story.append(styled_table([
        ["Total Tasks", "Completed", "In Progress", "Pending", "Overdue", "Completion %"],
        [str(kpis["total"]), str(kpis["completed"]), str(kpis["in_progress"]),
         str(kpis["pending"]), str(kpis["overdue"]), f"{kpis['completion_rate']}%"],
    ]))

    story.append(Paragraph("Sales Collection vs Target", h2))
    story.append(styled_table([
        ["Target Total", "Collected Total", "Achievement %", "Sales Items"],
        [f"Rs. {sales['target_total']:,.0f}", f"Rs. {sales['collected_total']:,.0f}",
         f"{sales['achievement_pct']}%", str(sales["count"])],
    ], header_bg=gold))

    for key, label in (("by_head", "Activity Head-wise Performance"), ("by_hq", "HQ-wise Performance"),
                       ("by_assignee", "Assignee-wise Performance"),
                       ("by_role", "Role-wise Performance"), ("by_frequency", "Frequency-wise Performance")):
        rows = data[key]
        if not rows:
            continue
        story.append(Paragraph(label, h2))
        table_rows = [["Name", "Total", "Completed", "In Progress", "Pending", "Overdue", "Rate"]]
        for r in rows:
            display = HEAD_LABELS.get(r["name"], r["name"]) if key == "by_head" else r["name"]
            table_rows.append([
                Paragraph(str(display), styles["Normal"]),
                str(r["total"]), str(r["completed"]), str(r["in_progress"]),
                str(r["pending"]), str(r["overdue"]), f"{r['completion_rate']}%",
            ])
        story.append(styled_table(table_rows, col_widths=[60 * mm, None, None, None, None, None, None]))

    # ---------- Performance data (latest month) ----------
    perf_months = await db.brand_performance.distinct("month")
    month_label_map = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                       7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

    def pm_label(m):
        y, mo = m.split("-")
        return f"{month_label_map[int(mo)]} {y}"

    if perf_months:
        pm = max(perf_months)
        brand_rows = await db.brand_performance.find({"month": pm}, {"_id": 0}).to_list(200)
        brand_rows.sort(key=lambda x: -(x.get("sales_total") or 0))
        if brand_rows:
            story.append(Paragraph(f"Brand Performance — {pm_label(pm)}", h2))
            table_rows = [["Brand", "Target", "Sales", "Ach %", "Top Territory", "Low Territory"]]
            for b in brand_rows:
                table_rows.append([
                    Paragraph(b["brand"], styles["Normal"]),
                    f"{b['target']:,.0f}" if b.get("target") else "-",
                    f"{b['sales_total']:,.0f}" if b.get("sales_total") is not None else "-",
                    f"{b['achievement_pct']}%" if b.get("achievement_pct") is not None else "-",
                    Paragraph(str(b.get("top_territory") or "-"), styles["Normal"]),
                    Paragraph(str(b.get("low_territory") or "-"), styles["Normal"]),
                ])
            story.append(styled_table(table_rows, col_widths=[38 * mm, 16 * mm, 16 * mm, 14 * mm, 45 * mm, 45 * mm], header_bg=gold))

        growth = await compute_brand_growth()
        if len(growth["months"]) >= 2:
            story.append(Paragraph(
                f"Brand Growth — Month over Month ({pm_label(growth['months'][0])} → {pm_label(growth['months'][-1])})", h2))
            table_rows = [["Brand"] + [pm_label(m) for m in growth["months"]] + ["Overall"]]
            for b in growth["brands"]:
                row_cells = [Paragraph(b["brand"], styles["Normal"])]
                for s in b["series"]:
                    cell_txt = f"{s['sales']:,.0f}" if s["sales"] is not None else "-"
                    if s["growth_pct"] is not None:
                        cell_txt += f" ({'+' if s['growth_pct'] >= 0 else ''}{s['growth_pct']}%)"
                    row_cells.append(cell_txt)
                overall = b["overall_growth_pct"]
                row_cells.append(f"{'+' if overall is not None and overall >= 0 else ''}{overall}%" if overall is not None else "-")
                table_rows.append(row_cells)
            story.append(styled_table(table_rows, col_widths=[40 * mm] + [None] * (len(growth["months"]) + 1)))

    terr_months = await db.territory_performance.distinct("month")
    if terr_months:
        tm = max(terr_months)
        terr_rows = await db.territory_performance.find({"month": tm}, {"_id": 0}).to_list(1000)
        regions = {}
        for it in terr_rows:
            regions.setdefault(it.get("region") or "Other", []).append(it)
        if regions:
            story.append(Paragraph(f"Territory Performance by Region — {pm_label(tm)}", h2))
            table_rows = [["Region", "HQs", "Target (L)", "Sales (L)", "Ach %"]]
            region_list = []
            for region, its in regions.items():
                tgt = sum(i.get("target") or 0 for i in its)
                sal = sum(i.get("sales_total") or 0 for i in its)
                region_list.append((region, len(its), tgt, sal))
            region_list.sort(key=lambda x: -x[3])
            for region, n, tgt, sal in region_list:
                table_rows.append([
                    Paragraph(region, styles["Normal"]), str(n),
                    f"{tgt:,.2f}", f"{sal:,.2f}",
                    f"{sal / tgt * 100:.0f}%" if tgt else "-",
                ])
            story.append(styled_table(table_rows, col_widths=[55 * mm, None, None, None, None]))

    mgmt_months = await db.management_dashboard.distinct("month")
    if mgmt_months:
        mm_ = max(mgmt_months)
        mdoc = await db.management_dashboard.find_one({"month": mm_}, {"_id": 0})
        metrics = (mdoc or {}).get("metrics") or {}
        numeric_labels = [("primary_sales", "Total Primary Sales (Lacs)"),
                          ("secondary_sales", "Total Secondary Sales (Lacs)"),
                          ("active_doctors", "Active Doctors"),
                          ("new_prescribers", "New Prescribers")]
        rows_num = []
        for key, label in numeric_labels:
            m = metrics.get(key)
            if not m:
                continue
            weeks = m.get("weeks") or [None] * 4
            rows_num.append([label] + [f"{w:,.2f}".rstrip("0").rstrip(".") if isinstance(w, (int, float)) else "-" for w in weeks])
        if rows_num:
            story.append(Paragraph(f"Management Dashboard — {pm_label(mm_)}", h2))
            story.append(styled_table([["Metric", "W1", "W2", "W3", "W4"]] + rows_num,
                                      col_widths=[60 * mm, None, None, None, None]))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    filename = f"surishi_{period}_report_{data['range']['start']}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/")
async def root():
    return {"message": "Surishi Pharma Marketing Execution API"}


# ------------------- Seed demo users -------------------
DEMO_USERS = [
    {"name": "Marketing Head", "email": "head@surishi.com", "role": "marketing_head"},
    {"name": "Deputy Marketing Head", "email": "deputy@surishi.com", "role": "marketing_deputy_head"},
    {"name": "Product Executive", "email": "pe@surishi.com", "role": "product_executive"},
    {"name": "General Manager", "email": "gm@surishi.com", "role": "general_manager"},
    {"name": "CEO", "email": "ceo@surishi.com", "role": "ceo"},
    {"name": "Chairman", "email": "chairman@surishi.com", "role": "chairman"},
    {"name": "AGM", "email": "agm@surishi.com", "role": "agm"},
    {"name": "Business Manager", "email": "bm@surishi.com", "role": "business_manager"},
]
DEMO_PASSWORD = "Surishi@123"


@app.on_event("startup")
async def seed_users():
    for u in DEMO_USERS:
        existing = await db.users.find_one({"email": u["email"]})
        if not existing:
            await db.users.insert_one({
                "id": str(uuid.uuid4()),
                "name": u["name"],
                "email": u["email"],
                "role": u["role"],
                "hq": None,
                "password_hash": pwd_hash.hash(DEMO_PASSWORD),
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
    logger.info("Demo users seeded")


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
